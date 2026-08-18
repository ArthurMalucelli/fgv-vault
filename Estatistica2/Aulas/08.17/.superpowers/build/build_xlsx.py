#!/usr/bin/env python3
"""Gera TestesHipotese.xlsx a partir de dados.json. Fórmulas vivas, sem valor colado.
Convenção: azul = input que o Arthur pode mudar, preto = fórmula. Fonte Arial."""
import json, os, re, copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
AULA = os.path.dirname(os.path.dirname(HERE))
d = json.load(open(os.path.join(HERE, "dados.json")))
cell_map = {}

XLFN = ["NORM.S.INV", "NORM.S.DIST", "T.INV.2T", "T.INV", "T.DIST.2T", "T.DIST.RT", "T.DIST",
        "CHISQ.INV.RT", "CHISQ.DIST.RT", "CHISQ.TEST", "T.TEST", "Z.TEST", "STDEV.S"]
def F(s):
    """prefixa _xlfn. nas funções pós-2007 (openpyxl não faz sozinho; sem isso o Excel dá #NAME?)"""
    for fn in sorted(XLFN, key=len, reverse=True):
        s = re.sub(r'(?<![\w.])' + re.escape(fn) + r'\(', '_xlfn.' + fn + '(', s)
    return s

FONT = "Arial"
BASE = Font(name=FONT, size=10)
BLUE = Font(name=FONT, size=10, color="0000FF")
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=13, bold=True)
SUB = Font(name=FONT, size=10, italic=True, color="555555")
GREY = PatternFill("solid", fgColor="EFEFEF")
LEGENDA = "Azul = input (pode mudar e ver a decisão virar). Preto = fórmula. "

def title(ws, text, sub=""):
    ws["A1"] = text; ws["A1"].font = TITLE
    ws["A2"] = LEGENDA + sub; ws["A2"].font = SUB

def put(ws, cell, value, *, inp=False, fmt=None, bold=False, name=None):
    c = ws[cell]
    c.value = F(value) if isinstance(value, str) and value.startswith("=") else value
    c.font = BLUE if inp else (BOLD if bold else BASE)
    if fmt: c.number_format = fmt
    if name: cell_map[name] = f"{ws.title}!{cell}"
    return c

def rows(ws, col_label, col_val, start, items):
    """items: lista de (label, value, kwargs). Escreve rótulo e valor linha a linha."""
    r = start
    for label, value, kw in items:
        lc = ws[f"{col_label}{r}"]; lc.value = label
        lc.font = BOLD if label.startswith(("1.", "2.", "3.", "4.", "5.", "Leitura")) else BASE
        if value is not None: put(ws, f"{col_val}{r}", value, **kw)
        r += 1

def widths(ws, spec):
    for col, w in spec.items(): ws.column_dimensions[col].width = w

def data_col(ws, col, start, header, values, fmt="0.0"):
    ws[f"{col}{start-1}"] = header; ws[f"{col}{start-1}"].font = BOLD
    for i, v in enumerate(values):
        put(ws, f"{col}{start+i}", v, inp=True, fmt=fmt)
    return f"{col}{start}:{col}{start+len(values)-1}"

wb = Workbook()

# ---------------- Mapa ----------------
ws = wb.active; ws.title = "Mapa"
title(ws, "Mapa de funções: Excel (EN / PT) x R", "Sintaxe EN usa vírgula entre argumentos; Excel em português usa ponto e vírgula.")
head = ["Objetivo", "Excel (EN)", "Excel (PT)", "R", "Devolve", "Quando usar / observação"]
for j, h in enumerate(head, 1):
    c = ws.cell(row=4, column=j, value=h); c.font = BOLD; c.fill = GREY
mapa = [
 ("Tamanho da amostra n", "COUNT(range)", "CONT.NÚM(intervalo)", "length(x)", "n", ""),
 ("Média amostral", "AVERAGE(range)", "MÉDIA(intervalo)", "mean(x)", "x̄", ""),
 ("Desvio-padrão amostral s", "STDEV.S(range)", "DESVPAD.A(intervalo)", "sd(x)", "s (divide por n-1)", "STDEV.P divide por n: errado pra amostra"),
 ("Raiz e valor absoluto", "SQRT(x), ABS(x)", "RAIZ(x), ABS(x)", "sqrt(x), abs(x)", "", ""),
 ("Z crítico unicaudal", "NORM.S.INV(1-α)", "INV.NORMP.N(1-α)", "qnorm(1-alpha)", "1,645 pra 5%", "cauda esquerda: NORM.S.INV(α) = -1,645"),
 ("Z crítico bicaudal", "NORM.S.INV(1-α/2)", "INV.NORMP.N(1-α/2)", "qnorm(1-alpha/2)", "1,960 pra 5%", "α/2 em cada cauda"),
 ("valor-p Z, cauda direita", "1-NORM.S.DIST(z,TRUE)", "1-DIST.NORMP.N(z;VERDADEIRO)", "1-pnorm(z)", "área à direita de z", "cauda esquerda: NORM.S.DIST(z,TRUE)"),
 ("valor-p Z, bicaudal", "2*(1-NORM.S.DIST(ABS(z),TRUE))", "2*(1-DIST.NORMP.N(ABS(z);VERDADEIRO))", "2*(1-pnorm(abs(z)))", "dobro da cauda", ""),
 ("t crítico unicaudal direita", "T.INV(1-α,gl)", "INV.T(1-α;gl)", "qt(1-alpha,gl)", "ex. 1,729 (gl 19)", "T.INV(α,gl) dá o da esquerda (negativo)"),
 ("t crítico bicaudal", "T.INV.2T(α,gl)", "INV.T.BC(α;gl)", "qt(1-alpha/2,gl)", "ex. 2,093 (gl 19)", "entra α inteiro, a função divide"),
 ("valor-p t, cauda direita", "T.DIST.RT(t,gl)", "DIST.T.CD(t;gl)", "pt(t,gl,lower.tail=FALSE)", "área à direita", ""),
 ("valor-p t, cauda esquerda", "T.DIST(t,gl,TRUE)", "DIST.T(t;gl;VERDADEIRO)", "pt(t,gl)", "área à esquerda", ""),
 ("valor-p t, bicaudal", "T.DIST.2T(ABS(t),gl)", "DIST.T.BC(ABS(t);gl)", "2*pt(-abs(t),gl)", "dobro da cauda", "T.DIST.2T exige t >= 0: use ABS"),
 ("Qui-quadrado crítico", "CHISQ.INV.RT(α,gl)", "INV.QUIQUA.CD(α;gl)", "qchisq(1-alpha,gl)", "ex. 5,991 (gl 2)", "sempre cauda direita"),
 ("valor-p qui-quadrado", "CHISQ.DIST.RT(x2,gl)", "DIST.QUIQUA.CD(x2;gl)", "pchisq(x2,gl,lower.tail=FALSE)", "área à direita", ""),
 ("Atalho: t pareado", "T.TEST(dep,ant,caudas,1)", "TESTE.T(dep;ant;caudas;1)", "t.test(dep,ant,paired=TRUE)", "só o valor-p", "caudas=1 devolve a cauda do |t|, não sabe a direção de H1"),
 ("Atalho: t duas amostras (Welch)", "T.TEST(A,B,2,3)", "TESTE.T(A;B;2;3)", "t.test(A,B)", "só o valor-p", "tipo 3 = variâncias diferentes; tipo 2 = iguais"),
 ("Atalho: Z uma média", "Z.TEST(range,μ0,σ)", "TESTE.Z(intervalo;μ0;σ)", "(na mão)", "valor-p da cauda superior", "cauda inferior: 1-Z.TEST; bicaudal: 2*MIN(Z.TEST,1-Z.TEST)"),
 ("Atalho: qui-quadrado", "CHISQ.TEST(obs,esp)", "TESTE.QUIQUA(obs;esp)", "chisq.test(x=obs,p=p) / chisq.test(tab)", "só o valor-p", "esperadas você calcula; a estatística não sai da função"),
 ("Proporção: função pronta", "(na mão)", "(na mão)", "prop.test(x,n,p=p0,alternative=,correct=FALSE)", "X-squared = z², mesmo valor-p", "default correct=TRUE muda o valor-p"),
]
for i, row in enumerate(mapa, 5):
    for j, v in enumerate(row, 1):
        ws.cell(row=i, column=j, value=v).font = BASE
widths(ws, {"A": 30, "B": 32, "C": 36, "D": 40, "E": 26, "F": 52})

# ---------------- 1_Z_Media ----------------
ws = wb.create_sheet("1_Z_Media"); v = d["t1_vendas"]
title(ws, "1. Z para uma média, σ conhecido: vendas diárias (meta R$ 500)", "Amostra de 36 dias, média 520, σ populacional 30. 'As vendas aumentaram?'")
rows(ws, "C", "D", 4, [
 ("1. H0", "μ ≤ 500", {}), ("1. H1", "μ > 500  (unicaudal à direita)", {}),
 ("2. α", v["alpha"], {"inp": True}), ("μ0 (H0)", v["mu0"], {"inp": True}),
 ("σ populacional (dado)", v["sigma"], {"inp": True}), ("n", v["n"], {"inp": True}), ("x̄ (média amostral)", v["xbar"], {"inp": True}),
 ("3. Erro-padrão σ/√n", "=D8/SQRT(D9)", {"fmt": "0.000"}),
 ("3. z", "=(D10-D7)/D11", {"fmt": "0.000", "name": "t1_z"}),
 ("4. z crítico (1-α)", "=NORM.S.INV(1-D6)", {"fmt": "0.000", "name": "t1_crit"}),
 ("4. valor-p (área à direita)", "=1-NORM.S.DIST(D12,TRUE)", {"fmt": "0.00000", "name": "t1_p"}),
 ("5. Decisão", '=IF(D14<D6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("5. Interpretação", "Há evidência, a 5%, de que a média de vendas superou a meta de 500. O 520 é da amostra, não é 'a' média populacional.", {}),
])
widths(ws, {"C": 30, "D": 60})

# ---------------- 2_T_Media (2a sumário | 2b dados brutos) ----------------
ws = wb.create_sheet("2_T_Media"); s2 = d["t2a_satisf"]
title(ws, "2. t para uma média (σ desconhecido): 2a satisfação (sumário) e 2b SLA 48h (dados brutos)")
ws["C3"] = "2a. Satisfação: benchmark 7,0, n = 20, x̄ = 6,8, s = 0,5"; ws["C3"].font = BOLD
rows(ws, "C", "D", 4, [
 ("2. α", s2["alpha"], {"inp": True}), ("μ0", s2["mu0"], {"inp": True}), ("n", s2["n"], {"inp": True}),
 ("x̄", s2["xbar"], {"inp": True}), ("s (amostral)", s2["s"], {"inp": True}),
 ("gl = n-1", "=D6-1", {}), ("3. Erro-padrão s/√n", "=D8/SQRT(D6)", {"fmt": "0.000"}),
 ("3. t", "=(D7-D5)/D10", {"fmt": "0.000", "name": "t2a_t"}),
 ("Leitura 1: H1 μ ≠ 7 (bicaudal)", None, {}),
 ("4. t crítico bicaudal ±", "=T.INV.2T(D4,D9)", {"fmt": "0.000", "name": "t2a_crit_bi"}),
 ("4. valor-p bicaudal", "=T.DIST.2T(ABS(D11),D9)", {"fmt": "0.0000", "name": "t2a_p_bi"}),
 ("5. Decisão bicaudal", '=IF(D14<D4,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Leitura 2: H1 μ < 7 (unicaudal esquerda)", None, {}),
 ("4. t crítico esquerda", "=T.INV(D4,D9)", {"fmt": "0.000", "name": "t2a_crit_esq"}),
 ("4. valor-p esquerda", "=T.DIST(D11,D9,TRUE)", {"fmt": "0.0000", "name": "t2a_p_esq"}),
 ("5. Decisão unicaudal", '=IF(D18<D4,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Pegadinha", "Mesma amostra, decisões diferentes: a cauda vem do enunciado.", {}),
])
rng_sla = data_col(ws, "A", 4, "SLA (h)", d["t2b_sla"])
ws["F3"] = "2b. SLA 48h: n = 25 entregas na coluna A. 'Está estourando o SLA?'"; ws["F3"].font = BOLD
rows(ws, "F", "G", 4, [
 ("1. H0", "μ ≤ 48", {}), ("1. H1", "μ > 48  (unicaudal à direita)", {}),
 ("2. α", 0.05, {"inp": True}), ("μ0", d["t2b_mu0"], {"inp": True}),
 ("3. n", f"=COUNT({rng_sla})", {}), ("3. x̄", f"=AVERAGE({rng_sla})", {"fmt": "0.000", "name": "t2b_xbar"}),
 ("3. s", f"=STDEV.S({rng_sla})", {"fmt": "0.000", "name": "t2b_s"}),
 ("3. Erro-padrão s/√n", "=G10/SQRT(G8)", {"fmt": "0.000"}), ("gl = n-1", "=G8-1", {}),
 ("3. t", "=(G9-G7)/G11", {"fmt": "0.000", "name": "t2b_t"}),
 ("4. t crítico direita", "=T.INV(1-G6,G12)", {"fmt": "0.000", "name": "t2b_crit"}),
 ("4. valor-p direita", "=T.DIST.RT(G13,G12)", {"fmt": "0.0000", "name": "t2b_p"}),
 ("5. Decisão a 5%", '=IF(G15<G6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("α alternativo", 0.01, {"inp": True}),
 ("5. Decisão a 1%", '=IF(G15<G17,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("5. Interpretação", "A 5% há evidência de que o tempo médio passa das 48h prometidas; a 1% a amostra não basta. Reportar os dois.", {}),
])
widths(ws, {"A": 10, "C": 34, "D": 22, "F": 30, "G": 40})

# ---------------- 3_Pareado ----------------
ws = wb.create_sheet("3_Pareado")
title(ws, "3. t pareado: vendas diárias (R$ mil) das mesmas 12 lojas, antes e depois da campanha", "D = Depois − Antes. 'A campanha aumentou as vendas?'")
n3 = len(d["t3_antes"])
r_ant = data_col(ws, "A", 4, "Antes", d["t3_antes"]); r_dep = data_col(ws, "B", 4, "Depois", d["t3_depois"])
ws["C3"] = "D = Depois-Antes"; ws["C3"].font = BOLD
for i in range(n3): put(ws, f"C{4+i}", f"=B{4+i}-A{4+i}", fmt="0.0")
r_D = f"C4:C{3+n3}"
rows(ws, "E", "F", 4, [
 ("1. H0", "μ_D ≤ 0", {}), ("1. H1", "μ_D > 0  (unicaudal à direita)", {}), ("2. α", 0.05, {"inp": True}),
 ("3. n (pares)", f"=COUNT({r_D})", {}), ("3. D médio", f"=AVERAGE({r_D})", {"fmt": "0.000", "name": "t3_dbar"}),
 ("3. s_D", f"=STDEV.S({r_D})", {"fmt": "0.000", "name": "t3_sd"}),
 ("3. Erro-padrão s_D/√n", "=F9/SQRT(F7)", {"fmt": "0.000"}), ("gl = n-1", "=F7-1", {}),
 ("3. t", "=(F8-0)/F10", {"fmt": "0.000", "name": "t3_t"}),
 ("4. t crítico direita", "=T.INV(1-F6,F11)", {"fmt": "0.000", "name": "t3_crit"}),
 ("4. valor-p direita", "=T.DIST.RT(F12,F11)", {"fmt": "0.0000", "name": "t3_p"}),
 ("5. Decisão", '=IF(F14<F6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("IC 95% de μ_D: inferior", "=F8-T.INV.2T(0.05,F11)*F10", {"fmt": "0.00", "name": "t3_ic_lo"}),
 ("IC 95% de μ_D: superior", "=F8+T.INV.2T(0.05,F11)*F10", {"fmt": "0.00", "name": "t3_ic_hi"}),
 ("Atalho T.TEST(dep,ant,1,1)", f"=T.TEST({r_dep},{r_ant},1,1)", {"fmt": "0.0000", "name": "t3_p_ttest"}),
 ("Obs. do atalho", "caudas=1 devolve a cauda de |t|: só vale se D médio tem o sinal de H1", {}),
 ("5. Interpretação", "Há evidência de que a campanha aumentou as vendas. Efeito médio = D médio, com o IC como faixa plausível.", {}),
])
widths(ws, {"A": 9, "B": 9, "C": 16, "E": 30, "F": 44})

# ---------------- 4_Welch ----------------
ws = wb.create_sheet("4_Welch")
title(ws, "4. t para duas médias independentes (Welch): ticket médio (R$) loja A vs loja B", "Amostras de clientes diferentes. 'Os tickets médios diferem?' Extensão além dos slides.")
r_A = data_col(ws, "A", 4, "Loja A", d["t4_A"]); r_B = data_col(ws, "B", 4, "Loja B", d["t4_B"])
rows(ws, "D", "E", 4, [
 ("1. H0", "μ_A = μ_B", {}), ("1. H1", "μ_A ≠ μ_B  (bicaudal)", {}), ("2. α", 0.05, {"inp": True}),
 ("nA", f"=COUNT({r_A})", {}), ("nB", f"=COUNT({r_B})", {}),
 ("x̄A", f"=AVERAGE({r_A})", {"fmt": "0.000"}), ("x̄B", f"=AVERAGE({r_B})", {"fmt": "0.000"}),
 ("sA", f"=STDEV.S({r_A})", {"fmt": "0.000"}), ("sB", f"=STDEV.S({r_B})", {"fmt": "0.000"}),
 ("3. Erro-padrão √(sA²/nA+sB²/nB)", "=SQRT(E11^2/E7+E12^2/E8)", {"fmt": "0.000"}),
 ("3. t", "=(E9-E10)/E13", {"fmt": "0.000", "name": "t4_t"}),
 ("gl de Welch", "=E13^4/((E11^2/E7)^2/(E7-1)+(E12^2/E8)^2/(E8-1))", {"fmt": "0.00", "name": "t4_gl"}),
 ("4. t crítico bicaudal ±", "=T.INV.2T(E6,E15)", {"fmt": "0.000", "name": "t4_crit"}),
 ("4. valor-p bicaudal", "=T.DIST.2T(ABS(E14),E15)", {"fmt": "0.0000", "name": "t4_p"}),
 ("5. Decisão", '=IF(E17<E6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Atalho T.TEST(A,B,2,3)", f"=T.TEST({r_A},{r_B},2,3)", {"fmt": "0.0000", "name": "t4_p_ttest"}),
 ("Obs.", "Excel trunca gl não inteiro em T.DIST/T.INV; R usa gl fracionário. Diferença na 3ª casa do valor-p é isso.", {}),
 ("5. Interpretação", "A 5% a amostra não sustenta que os tickets médios diferem (valor-p acima de α por pouco). Não é prova de igualdade.", {}),
])
widths(ws, {"A": 9, "B": 9, "D": 34, "E": 44})

# ---------------- 5_Proporcao ----------------
ws = wb.create_sheet("5_Proporcao"); g = d["t5_golfe"]
title(ws, "5. Z para uma proporção: clube de golfe, baseline 20% de mulheres", "Após a promoção: 100 mulheres em 400 jogadores. 'A proporção aumentou?'")
rows(ws, "C", "D", 4, [
 ("1. H0", "p ≤ 0,20", {}), ("1. H1", "p > 0,20  (unicaudal à direita)", {}), ("2. α", g["alpha"], {"inp": True}),
 ("x (sucessos)", g["x"], {"inp": True}), ("n", g["n"], {"inp": True}), ("p0 (H0)", g["p0"], {"inp": True}),
 ("3. p̂ = x/n", "=D7/D8", {"fmt": "0.000"}),
 ("3. Erro-padrão √(p0(1-p0)/n)", "=SQRT(D9*(1-D9)/D8)", {"fmt": "0.0000"}),
 ("3. z", "=(D10-D9)/D11", {"fmt": "0.000", "name": "t5_z"}),
 ("4. z crítico (1-α)", "=NORM.S.INV(1-D6)", {"fmt": "0.000", "name": "t5_crit"}),
 ("4. valor-p direita", "=1-NORM.S.DIST(D12,TRUE)", {"fmt": "0.0000", "name": "t5_p"}),
 ("5. Decisão", '=IF(D14<D6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Erro comum: EP com p̂ no denominador", "=SQRT(D10*(1-D10)/D8)", {"fmt": "0.0000"}),
 ("z errado com esse EP", "=(D10-D9)/D16", {"fmt": "0.000"}),
 ("5. Interpretação", "Fortes evidências de que a proporção de mulheres passou de 20%. Não diz que 'é 25%': 0,25 é da amostra.", {}),
])
widths(ws, {"C": 36, "D": 60})

# ---------------- 6_Qui_Aderencia ----------------
ws = wb.create_sheet("6_Qui_Aderencia"); q6 = d["t6_pagto"]; k = len(q6["obs"])
title(ws, "6. Qui-quadrado de aderência: mix de pagamento (esperado 45/35/20)", "n = 200 compras. 'A distribuição observada segue o mix esperado?'")
for j, h in enumerate(["Categoria", "Observado O", "p esperada", "Esperado E = n·p", "(O-E)²/E"], 1):
    c = ws.cell(row=3, column=j, value=h); c.font = BOLD; c.fill = GREY
for i in range(k):
    r = 4 + i
    ws[f"A{r}"] = q6["cat"][i]; ws[f"A{r}"].font = BASE
    put(ws, f"B{r}", q6["obs"][i], inp=True); put(ws, f"C{r}", q6["p"][i], inp=True)
    put(ws, f"D{r}", f"=C{r}*$B${4+k}", fmt="0.00"); put(ws, f"E{r}", f"=(B{r}-D{r})^2/D{r}", fmt="0.0000")
tot = 4 + k
ws[f"A{tot}"] = "Total"; ws[f"A{tot}"].font = BOLD
put(ws, f"B{tot}", f"=SUM(B4:B{tot-1})"); put(ws, f"C{tot}", f"=SUM(C4:C{tot-1})"); put(ws, f"D{tot}", f"=SUM(D4:D{tot-1})")
rows(ws, "G", "H", 4, [
 ("1. H0", "distribuição observada segue 45/35/20", {}), ("1. H1", "pelo menos uma proporção difere", {}), ("2. α", 0.05, {"inp": True}),
 ("3. χ² = Σ(O-E)²/E", f"=SUM(E4:E{tot-1})", {"fmt": "0.000", "name": "t6_x2"}),
 ("gl = k-1", f"=COUNT(B4:B{tot-1})-1", {}),
 ("4. χ² crítico (cauda direita)", "=CHISQ.INV.RT(H6,H8)", {"fmt": "0.000", "name": "t6_crit"}),
 ("4. valor-p", "=CHISQ.DIST.RT(H7,H8)", {"fmt": "0.0000", "name": "t6_p"}),
 ("5. Decisão", '=IF(H10<H6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Atalho CHISQ.TEST(obs,esp)", f"=CHISQ.TEST(B4:B{tot-1},D4:D{tot-1})", {"fmt": "0.0000", "name": "t6_p_chisqtest"}),
 ("5. Interpretação", "A amostra é consistente com o mix esperado; não há evidência de mudança no comportamento de pagamento.", {}),
])
widths(ws, {"A": 12, "B": 13, "C": 12, "D": 18, "E": 12, "G": 32, "H": 46})

# ---------------- 7_Qui_Independencia ----------------
ws = wb.create_sheet("7_Qui_Independencia"); q7 = d["t7_canal"]; R, C = len(q7["linhas"]), len(q7["colunas"])
title(ws, "7. Qui-quadrado de independência: canal de marketing x compra (n = 320 leads)", "'A chance de compra depende do canal?'")
def matriz(ws, top, left, label, cells, fmt=None, inp=False):
    ws.cell(row=top, column=left, value=label).font = BOLD
    for j, h in enumerate(q7["colunas"]): ws.cell(row=top, column=left+1+j, value=h).font = BOLD
    for i, rl in enumerate(q7["linhas"]):
        ws.cell(row=top+1+i, column=left, value=rl).font = BASE
        for j in range(C):
            put(ws, f"{get_column_letter(left+1+j)}{top+1+i}", cells[i][j], inp=inp, fmt=fmt)
    return top+1, left+1   # primeira célula de dados (linha, coluna)
r0, c0 = matriz(ws, 3, 1, "Observado", q7["obs"], inp=True)              # B4:C6
ws.cell(row=3, column=4, value="Total linha").font = BOLD
for i in range(R): put(ws, f"D{r0+i}", f"=SUM(B{r0+i}:C{r0+i})")
ws.cell(row=r0+R, column=1, value="Total coluna").font = BOLD
for j in range(C): put(ws, f"{get_column_letter(c0+j)}{r0+R}", f"=SUM({get_column_letter(c0+j)}{r0}:{get_column_letter(c0+j)}{r0+R-1})")
put(ws, f"D{r0+R}", f"=SUM(D{r0}:D{r0+R-1})")   # n em D7
nref = f"$D${r0+R}"
esp = [[f"=$D${r0+i}*{get_column_letter(c0+j)}${r0+R}/{nref}" for j in range(C)] for i in range(R)]
matriz(ws, 9, 1, "Esperado E", esp, fmt="0.00")                            # B10:C12
contrib = [[f"=({get_column_letter(c0+j)}{r0+i}-{get_column_letter(c0+j)}{10+i})^2/{get_column_letter(c0+j)}{10+i}" for j in range(C)] for i in range(R)]
matriz(ws, 15, 1, "(O-E)²/E", contrib, fmt="0.0000")                       # B16:C18
resid = [[f"=({get_column_letter(c0+j)}{r0+i}-{get_column_letter(c0+j)}{10+i})/SQRT({get_column_letter(c0+j)}{10+i})" for j in range(C)] for i in range(R)]
matriz(ws, 21, 1, "Resíduo (O-E)/√E", resid, fmt="0.00")                   # B22:C24
ws["A26"] = "Taxa de compra por canal"; ws["A26"].font = BOLD
for i in range(R):
    ws[f"A{27+i}"] = q7["linhas"][i]; ws[f"A{27+i}"].font = BASE
    put(ws, f"B{27+i}", f"=B{r0+i}/D{r0+i}", fmt="0%")
rows(ws, "F", "G", 4, [
 ("1. H0", "canal e compra são independentes", {}), ("1. H1", "existe associação", {}), ("2. α", 0.05, {"inp": True}),
 ("3. χ² = ΣΣ(O-E)²/E", "=SUM(B16:C18)", {"fmt": "0.000", "name": "t7_x2"}),
 ("gl = (r-1)(c-1)", "=(ROWS(B4:C6)-1)*(COLUMNS(B4:C6)-1)", {}),
 ("4. χ² crítico", "=CHISQ.INV.RT(G6,G8)", {"fmt": "0.000", "name": "t7_crit"}),
 ("4. valor-p", "=CHISQ.DIST.RT(G7,G8)", {"fmt": "0.0000", "name": "t7_p"}),
 ("5. Decisão", '=IF(G10<G6,"Rejeita H0","Não rejeita H0")', {"bold": True}),
 ("Atalho CHISQ.TEST(obs,esp)", "=CHISQ.TEST(B4:C6,B10:C12)", {"fmt": "0.0000", "name": "t7_p_chisqtest"}),
 ("Menor esperada (regra ≥ 5)", "=MIN(B10:C12)", {"fmt": "0.00"}),
 ("5. Interpretação", "A probabilidade de compra depende do canal. Pelos resíduos e pelas taxas, Search é o canal forte, Social o fraco.", {}),
])
widths(ws, {"A": 22, "B": 13, "C": 13, "D": 12, "F": 30, "G": 52})

# ---------------- Exercicios ----------------
ws = wb.create_sheet("Exercicios")
title(ws, "Exercícios E1 a E4: só dados. Gabarito no fim do md.", "Monte você o bloco de passos ao lado de cada dataset.")
ws["A3"] = "E1 Atendimento (min). 'Mudou de 10 min?' α = 5%"; ws["A3"].font = BOLD
data_col(ws, "A", 5, "min", d["e1_atend"])
ws["C3"] = "E2 Conversão. Era 30%. 'Mudou?'"; ws["C3"].font = BOLD
rows(ws, "C", "D", 4, [("x (conversões)", d["e2_conv"]["x"], {"inp": True}), ("n (visitas)", d["e2_conv"]["n"], {"inp": True}), ("p0", d["e2_conv"]["p0"], {"inp": True})])
ws["F3"] = "E3a Mesmas 10 pessoas, min/tarefa antes e depois do treinamento. 'Reduziu?'"; ws["F3"].font = BOLD
data_col(ws, "F", 5, "Antes", d["e3a_antes"]); data_col(ws, "G", 5, "Depois", d["e3a_depois"])
ws["I3"] = "E3b Notas de duas turmas diferentes. 'As médias diferem?'"; ws["I3"].font = BOLD
data_col(ws, "I", 5, "Turma A", d["e3b_A"]); data_col(ws, "J", 5, "Turma B", d["e3b_B"])
ws["L3"] = "E4 Sabores. 'Preferência uniforme (25% cada)?'"; ws["L3"].font = BOLD
ws["L4"] = "Sabor"; ws["M4"] = "Observado"; ws["L4"].font = BOLD; ws["M4"].font = BOLD
for i, (c_, o) in enumerate(zip(d["e4_sabores"]["cat"], d["e4_sabores"]["obs"])):
    ws[f"L{5+i}"] = c_; ws[f"L{5+i}"].font = BASE; put(ws, f"M{5+i}", o, inp=True)
widths(ws, {"A": 10, "C": 18, "D": 10, "F": 10, "G": 10, "I": 10, "J": 10, "L": 10, "M": 12})

# fonte Arial em tudo, wrap nos textos longos
for w in wb.worksheets:
    for row in w.iter_rows():
        for c in row:
            if c.value is None: continue
            if c.font.name != FONT:
                f = copy.copy(c.font); c.font = Font(name=FONT, size=f.size or 10, bold=f.bold, italic=f.italic, color=f.color)
            if isinstance(c.value, str) and len(c.value) > 40 and not c.value.startswith("="):
                c.alignment = Alignment(wrap_text=True, vertical="top")

out = os.path.join(AULA, "TestesHipotese.xlsx")
wb.save(out)
json.dump(cell_map, open(os.path.join(HERE, "cell_map.json"), "w"), indent=1)
print("salvo:", out); print("cell_map:", len(cell_map), "entradas")
